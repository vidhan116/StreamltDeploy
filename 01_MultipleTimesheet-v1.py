##Imporlit run Multiting Libraries
import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime
import os
import glob
st.header("Timesheet processing Automation platform")
st.markdown("<h1 style='font-size: 30px; color: blue;'>Convert Apexon client employee time sheets(excel, csv file) to Nest standard format.</h1>", unsafe_allow_html=True)
# st.title("*Convert Apexon client employee time sheets(excel, csv file) to Nest standard format.*")
##++++++++++++++++++++++++++++++++++++Creating Folder structures++++++++++++++++++++++++++++++++++++++++++++++++

Output_FolderPath= 'OutputResultFolder'
Individual_FolderPath= 'IndividualEmpFileFolder'
Log_FolderPath= 'LogFolder'
Empmaster_FolderPath= 'Employee master'
Input_dataFolder= 'InputDataFolder'
# Create the directory if it doesn't exist
if not os.path.exists(Output_FolderPath):
    os.makedirs(Output_FolderPath)
if not os.path.exists(Individual_FolderPath):
    os.makedirs(Individual_FolderPath)
if not os.path.exists(Log_FolderPath):
    os.makedirs(Log_FolderPath)
if not os.path.exists(Empmaster_FolderPath):
    os.makedirs(Empmaster_FolderPath)
if not os.path.exists(Input_dataFolder):
    os.makedirs(Input_dataFolder)

Reset_files=st.sidebar.button("Click here to clear folder")
if Reset_files:
    # +++++++++++++++++++++++++++++++++++++++++++Deleting previous files from Output folder++++++++++++++++++++++++++++++++++++++++++++

    def delete_OutputFolder_files(folder_path, file_list):
        for file_name in file_list:
            file_path = os.path.join(folder_path, file_name)
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"Deleted: {file_path}")
            else:
                print(f"File not found: {file_path}")


    # Example usage
    folder_path = 'OutputResultFolder'
    file_list = ['AirBus_result.csv', 'Ford_result.csv', 'ICON_result.csv']
    delete_OutputFolder_files(folder_path, file_list)


    # +++++++++++++++++++++++++++++++++++++++++++++++++++++++Deleting previous files from Log folder++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    def delete_LogFolder_files(folder_path, file_list):
        for file_name in file_list:
            file_path = os.path.join(folder_path, file_name)
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"Deleted: {file_path}")
            else:
                print(f"File not found: {file_path}")


    # Example usage
    folder_path = 'LogFolder'
    file_list = ['AirBusLog_file.csv', 'ICONLog_file.csv', 'Ford_logFile.csv']
    delete_LogFolder_files(folder_path, file_list)


    # ++++++++++++++++++++++++++++++++++++++++++++++++++++++++Deleting previous files from Individual folder+++++++++++++++++++++++++++++++++++++++++++++++++++++++

    def delete_Individual_files(folder_path):
        # Patterns for CSV and Excel files
        patterns = ["*.csv", "*.xlsx", "*.xls"]

        for pattern in patterns:
            # Get a list of all files matching the pattern
            files = glob.glob(os.path.join(folder_path, pattern))

            for file_path in files:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    print(f"Deleted: {file_path}")
                else:
                    print(f"File not found: {file_path}")


    # Example usage
    folder_path = 'IndividualEmpFileFolder'
    delete_Individual_files(folder_path)
    st.sidebar.success("Folder has been cleared")

    # ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++Process Starts+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

# Selecting Client timesheet
# st.header("Apexon projects employee timesheet conversion automation platform.")
# st.title("Convert Apexon client employee time sheets(excel, csv file) to Nest standard format.")
# Ind_Emp= st.button(label="Download project individual Employee file")

#Creating folder


##+++++++++++++++++++++++++++++++++++++++++++++++++AirBus process starts++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
## AIRBUS CLIENT TIMESHEET

# Define the paths for bridge files and log file
airbus_bridge_file = 'Employee master/AirBus bridge file.xlsx'


# Sidebar select box for project selection
project = st.selectbox("Select Project",("--Select--","AirBus", "FORD", "ICON"))
project_bridge_file= st.file_uploader("Upload project Bridge file")

#Define Path
log_file_path = 'LogFolder/AirBusLog_file.csv'
airbus_bridge_file = project_bridge_file

# Initialize log file
def initialize_log():
    with open(log_file_path, 'w') as log_file:
        log_file.write('Log File\n\n')


# Function to update log file
def update_log(message):
    with open(log_file_path, 'a') as log_file:
        log_file.write(message + '\n')


# Function to process timesheets

def process_timesheets(bridge_df, project_name, input_file):
    result_df = pd.DataFrame(
        columns=['SlNo', 'Project Name', 'Date_of_Work', 'Hours', 'Description', 'Is_Billable', 'Unique Employee ID'])

    for uploaded_file in input_file:
        if uploaded_file is not None:
            workbook = openpyxl.load_workbook(BytesIO(uploaded_file.read()), data_only=True)
            timesheet_found = False
            employee_found = False

            for sheetname in workbook.sheetnames:
                sheet = workbook[sheetname]

                if sheet['B6'].value == 'Name':
                    timesheet_found = True
                    employee_name = sheet['D6'].value

                    employee_row = bridge_df[bridge_df['EmployeeName'] == employee_name]
                    if not employee_row.empty:
                        employee_found = True
                        emp_id = employee_row['EmpID'].values[0]

                        if sheet['F11'].value == 'Date' and sheet['G11'].value == 'Hours':
                            temp_df = pd.DataFrame(
                                columns=['SlNo', 'Project Name', 'Date_of_Work', 'Hours', 'Description', 'Is_Billable',
                                         'Unique Employee ID'])
                            for row in range(12, 43):
                                work_date = sheet[f'F{row}'].value
                                hours = sheet[f'G{row}'].value
                                if work_date and hours:
                                    temp_df = pd.concat([temp_df, pd.DataFrame([{
                                        'SlNo': len(result_df) + len(temp_df) + 1,
                                        'Project Name': project_name,
                                        'Date_of_Work': work_date,
                                        'Hours': hours,
                                        'Description': 'Approved',
                                        'Is_Billable': 'YES',
                                        'Unique Employee ID': emp_id
                                    }])], ignore_index=True)
                            result_df = pd.concat([result_df, temp_df], ignore_index=True)
                        else:
                            update_log(f'{uploaded_file.name}: F11 or G11 does not contain the expected headers.')
                    else:
                        update_log(
                            f'{uploaded_file.name}: Employee {employee_name} not found in {project_name} bridge file.')

                    break

            if not timesheet_found:
                update_log(f'{uploaded_file.name}: Timesheet not found.')
            elif not employee_found:
                update_log(f'{uploaded_file.name}: Employee not found.')

    return result_df


# Main function to execute the app
def main():
    initialize_log()
    if project == "AirBus" and airbus_bridge_file:
        input_file = st.file_uploader("Upload AirBus project employee timesheet files.", type='XLSX', accept_multiple_files=True)
        airbus_bridge_df = pd.read_excel(airbus_bridge_file)
        result_df = process_timesheets(airbus_bridge_df, "Airbus", input_file)

        # Save result to CSV
        result_csv_path = 'OutputResultFolder/AirBus_result.csv'
        result_df.to_csv(result_csv_path, encoding= 'utf-8', index=False)

        # Display results and download link
        # generate_result= st.button('Click for AirBus output result')
        # if generate_result:
        #     st.write(result_df)
        AirBus_result_download= st.download_button(
            label="Download AirBus result file",
            data=open(result_csv_path, 'rb').read(),
            file_name='AirBus_result.csv',
            mime='text/csv'
        )
        if AirBus_result_download:
            st.success("AirBus result file has been downloaded")


        # Display log and download link
        AirBus_log_download=st.download_button(
            label="Download AirBus log file",
            data=open(log_file_path, 'rb').read(),
            file_name='AirBusLog_file.csv',
            mime='text/csv'
        )
        if AirBus_log_download:
            st.success("AirBus log file has been downloaded")


##++++++++++++++++++++++++++++++++++++++++++++++++++++++++Creating AirBus Individual Files++++++++++++++++++++++++++++++++++++++++
        # Ind_Emp = st.button("Download AirBus Individual Employee file")
        if project == 'AirBus':
            Ind_Emp = st.button("Download AirBus individual employee file")
            if Ind_Emp:
                def read_file(file_path):
                    try:
                        if file_path.endswith('.csv'):
                            df = pd.read_csv(file_path)
                        elif file_path.endswith('.xlsx'):
                            df = pd.read_excel(file_path)
                        else:
                            raise ValueError("Unsupported file format. Only CSV and Excel files are supported.")
                        return df
                    except Exception as e:
                        print(f"Error reading the file: {e}")
                        return None

                def write_files(df, output_dir):
                    try:
                        unique_employee_ids = df['Unique Employee ID'].unique()

                        if not os.path.exists(output_dir):
                            os.makedirs(output_dir)

                        for employee_id in unique_employee_ids:
                            employee_df = df[df['Unique Employee ID'] == employee_id]
                            output_file = os.path.join(output_dir, f"{employee_id}.csv")
                            employee_df.to_csv(output_file, index=False)

                        print(f"Files successfully written to {output_dir}")
                    except Exception as e:
                        print(f"Error writing the files: {e}")

                def main(input_file, output_dir):
                    df = read_file(input_file)
                    if df is not None:
                        write_files(df, output_dir)

                if __name__ == "__main__":
                    input_file = 'OutputResultFolder/AirBus_result.csv'  # Replace with the path to your input file
                    output_dir = 'IndividualEmpFileFolder'  # Replace with the path to your output directory
                    main(input_file, output_dir)


    else:
        pass
        # st.write("Please select a project & upload timesheet files.")


if __name__ == "__main__":
    main()


##+++++++++++++++++++++++++++++++++++++++++++++++++++++++AirBus Process Ends++++++++++++++++++++++++++++++++++++++++++++++++++++++



#++++++++++++++++++++++++++++++++++++++++++++ICON process starts++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

## ICON CLIENT TIMESHEET

# Define the paths for bridge files and log file
# ICON_bridge_file = 'Employee master/ICON bridge file.xlsx'
ICON_bridge_file= project_bridge_file
log_file_path = 'LogFolder/ICONLog_file.csv'


# Initialize log file
def initialize_log():
    with open(log_file_path, 'w') as log_file:
        log_file.write('Log File\n\n')


# Function to update log file
def update_log(message):
    with open(log_file_path, 'a') as log_file:
        log_file.write(message + '\n')


# Function to process timesheets

def process_timesheets(bridge_df, project_name, input_file):
    result_df = pd.DataFrame(
        columns=['SlNo', 'Project Name', 'Date_of_Work', 'Hours', 'Description', 'Is_Billable', 'Unique Employee ID'])

    for uploaded_file in input_file:
        if uploaded_file is not None:
            workbook = openpyxl.load_workbook(BytesIO(uploaded_file.read()), data_only=True)
            timesheet_found = False
            employee_found = False

            for sheetname in workbook.sheetnames:
                sheet = workbook[sheetname]

                if sheet['C1'].value == 'Date':
                    timesheet_found = True
                    employee_name = sheet['A2'].value
                    if employee_name is not None and employee_name !="":
                        employee_name=employee_name.upper()


                    employee_row = bridge_df[bridge_df['EmployeeName'] == employee_name]
                    if not employee_row.empty:
                        employee_found = True
                        emp_id = employee_row['EmpID'].values[0]

                        if sheet['C1'].value == 'Date' and sheet['D1'].value == 'Status':
                            temp_df = pd.DataFrame(
                                columns=['SlNo', 'Project Name', 'Date_of_Work', 'Hours', 'Description', 'Is_Billable',
                                         'Unique Employee ID'])
                            for row in range(2, 32):
                                work_date = sheet[f'C{row}'].value
                                status = sheet[f'D{row}'].value
                                if status== 'Present':
                                    temp_df = pd.concat([temp_df, pd.DataFrame([{
                                        'SlNo': len(result_df) + len(temp_df) + 1,
                                        'Project Name': project_name,
                                        'Date_of_Work': work_date,
                                        'Hours': 8,
                                        'Description': 'Approved',
                                        'Is_Billable': 'YES',
                                        'Unique Employee ID': emp_id
                                    }])], ignore_index=True)
                            result_df = pd.concat([result_df, temp_df], ignore_index=True)
                        else:
                            update_log(f'{uploaded_file.name}: C1 or D1 does not contain the expected headers.')
                    else:
                        update_log(
                            f'{uploaded_file.name}: Employee {employee_name} not found in {project_name} bridge file.')

                    break

            if not timesheet_found:
                update_log(f'{uploaded_file.name}: Timesheet not found.')
            elif not employee_found:
                update_log(f'{uploaded_file.name}: Employee not found.')

    return result_df


# Main function to execute the app
def main():
    initialize_log()
    if project == "ICON" and ICON_bridge_file:

        input_file = st.file_uploader("Upload ICON project employee timesheet files.", type='XLSX', accept_multiple_files=True)
        ICON_bridge_df = pd.read_excel(ICON_bridge_file)
        result_df = process_timesheets(ICON_bridge_df, "ICON", input_file)

        # Save result to CSV
        # result_csv_path = 'OutputResultFolder/ICON_result.csv'
        result_csv_path = 'OutputResultFolder/ICON_result.csv'
        # os.makedirs(output_dir, exist_ok=True)
        result_df.to_csv(result_csv_path, encoding= 'utf-8', index=False)

        # Display results and download link
        # generate_result= st.button('Click here for ICON output result')
        # if generate_result:
        #     st.write(result_df)
        ICON_result_download= st.download_button(
            label="Download ICON result file",
            data=open(result_csv_path, 'rb').read(),
            file_name='ICON_result.csv',
            mime='text/csv'
        )
        if ICON_result_download:
            st.success("ICON result file has been downloaded")


        # Display log and download link
        ICON_log_Download=st.download_button(
            label="Download ICON log file",
            data=open(log_file_path, 'rb').read(),
            file_name='ICONLog_file.csv',
            mime='text/csv'
        )
        if ICON_log_Download:
            st.success("ICON log file has been downloaded")


##++++++++++++++++++++++++++++++++++++++++++++++Creating ICON Individual file++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Ind_Emp = st.button("Download ICON Individual Employee file")
    if project== 'ICON':
        Ind_Emp = st.button("Download ICON individual Employee file")
        if Ind_Emp:
            def read_file(file_path):
                try:
                    if file_path.endswith('.csv'):
                        df = pd.read_csv(file_path)
                    elif file_path.endswith('.xlsx'):
                        df = pd.read_excel(file_path)
                    else:
                        raise ValueError("Unsupported file format. Only CSV and Excel files are supported.")
                    return df
                except Exception as e:
                    print(f"Error reading the file: {e}")
                    return None

            def write_files(df, output_dir):
                try:
                    unique_employee_ids = df['Unique Employee ID'].unique()

                    if not os.path.exists(output_dir):
                        os.makedirs(output_dir)

                    for employee_id in unique_employee_ids:
                        employee_df = df[df['Unique Employee ID'] == employee_id]
                        output_file = os.path.join(output_dir, f"{employee_id}.csv")
                        employee_df.to_csv(output_file, index=False)

                    print(f"Files successfully written to {output_dir}")
                except Exception as e:
                    print(f"Error writing the files: {e}")

            def main(input_file, output_dir):
                df = read_file(input_file)
                if df is not None:
                    write_files(df, output_dir)

            if __name__ == "__main__":
                input_file = 'OutputResultFolder/ICON_result.csv'  # Replace with the path to your input file
                output_dir = 'IndividualEmpFileFolder'  # Replace with the path to your output directory
                main(input_file, output_dir)


#+++++++++++++++++++++++++++++++++++++++++++++++++++ICON process Ends++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
if __name__ == "__main__":
    main()

##+++++++++++++++++++++++++++++++++++++++++++++++++++++++Ford Process Starts++++++++++++++++++++++++++++++++++++++++++++++++++++++

if project== 'FORD':
    timesheet_files = st.file_uploader("Upload Timesheet Files", type=["csv", "xlsx"], accept_multiple_files=True)

    if project_bridge_file and timesheet_files:
        try:
            ford_bridge_df = pd.read_excel(project_bridge_file)
            result_df_list = []
            log_result_df_list= []

            for file in timesheet_files:
                df = pd.read_csv(file, skiprows=1)
                if df.columns[0] == 'PO Number':
                    merged_df = pd.merge(df, ford_bridge_df[['File Id', 'EmpID']], on='File Id', how='inner')
                    filtered_df = merged_df[
                        (merged_df['Work Task'] != 'NBILL') &
                        (merged_df['Work Date'].notna()) &
                        (merged_df['Hours'].notna()) &
                        (merged_df['Hours'] != 0) &
                        (merged_df['Timesheet Status'] != 'MISSING')

                        ]
                    result_df = pd.DataFrame({
                        'SlNo': range(1, len(filtered_df) + 1),
                        'Project Name': 'FORD MAGNIT',
                        'Date_of_Work': filtered_df['Work Date'],
                        'Hours': filtered_df['Hours'],
                        'Description': filtered_df['Timesheet Status'],
                        'Is_Billable': 'YES',
                        'Unique Employee ID': filtered_df['EmpID']
                    })
                    result_df_list.append(result_df)

                    if result_df_list:
                        final_result_df = pd.concat(result_df_list)
                        result_csv_path = 'OutputResultFolder\Ford_result.csv'
                        final_result_df.to_csv(result_csv_path, index=False)
                        # st.write(final_result_df)
                        # Download Ford result file
            Ford_result_Download=st.download_button("Download Ford result file", open(result_csv_path, 'rb').read(),
                               file_name='Ford_result.csv', mime='text/csv')
            if Ford_result_Download:
                st.success("Ford result file has been downloaded")


            # Download Ford Individual file
            if st.button("Download Ford individual employee file"):
                if not final_result_df.empty:
                    output_dir = 'IndividualEmpFileFolder'
                    os.makedirs(output_dir, exist_ok=True)
                    for emp_id in final_result_df['Unique Employee ID'].unique():
                        emp_df = final_result_df[final_result_df['Unique Employee ID'] == emp_id]
                        emp_file_path = os.path.join(output_dir, f"{emp_id}.csv")
                        emp_df.to_csv(emp_file_path, index=False)
                    st.success(f"Individual employee files created in {output_dir}")
##++++++++++++++++++++++++++++++++++++++++++++++++++++++FORD Log File+++++++++++++++++++++++++++++++++++++++++++
                if df.columns[0] == 'PO Number':
                    log_merged_df = pd.merge(df, ford_bridge_df[['File Id', 'EmpID']], on='File Id', how='inner')
                    logged_df = log_merged_df[
                        (log_merged_df['Work Task'] == 'NBILL') |
                        (log_merged_df['Work Date'].isna()) |
                        (log_merged_df['Hours'].isna()) |
                        (log_merged_df['Hours'] == 0) |
                        (log_merged_df['Timesheet Status'] == 'MISSING')]

                    logresult_df = pd.DataFrame({
                        'SlNo': range(1, len(logged_df) + 1),
                        'Project Name': 'FORD MAGNIT',
                        'Date_of_Work': logged_df['Work Date'],
                        'Hours': logged_df['Hours'],
                        'Description': logged_df['Timesheet Status'],
                        'Is_Billable': 'YES',
                        'Unique Employee ID': logged_df['EmpID'],
                        'Work Task':logged_df['Work Task']
                    })
                    log_result_df_list.append(logresult_df)
                    if log_result_df_list:
                        logfinal_result_df = pd.concat(log_result_df_list)
                        logresult_csv_path = 'LogFolder\Ford_logFile.csv'
                        logfinal_result_df.to_csv(logresult_csv_path, index=False)
                        # st.write(logfinal_result_df)
                        st.download_button("Download Ford log file",
                                           open(logresult_csv_path, 'rb').read(),
                                           file_name='Ford_logFile.csv', mime='text/csv')
                        # st.success("Ford log file has been downloaded")


        except Exception as e:
            ("Processed Finished")

    ##+++++++++++++++++++++++++++++++++++++++++++++++++++++++Ford Process Ends++++++++++++++++++++++++++++++++++++++++++++++++++++++

    # if __name__ == "__main__":
    #     main()






