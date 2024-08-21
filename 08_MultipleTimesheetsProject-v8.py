import streamlit as st
import firebase_admin
from firebase_admin import credentials
from firebase_admin import auth
import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime
import os
import glob
from firebase_admin._auth_utils import UserNotFoundError
import firebase_admin
import pyrebase
import io
# Get current timestamp
timestamp = datetime.now().strftime("%d%m%Y_%H%M%S")

##+++++++++++++++++++++++++++++++++Set Background color++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# Custom CSS to set the background color to yellow
page_bg = """
           <style>
           .reportview-container {
               background: #87CEEB;
           }
           .main .block-container {
               background: #87CEEB;
           }
           </style>
           """

# Inject the CSS into the Streamlit app
st.markdown(page_bg, unsafe_allow_html=True)

##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# Initialize Firebase app
if not firebase_admin._apps:
    cred = credentials.Certificate('timesheet-conversion-platform-79b3689b08bb.json')
    firebase_admin.initialize_app(cred)

# Pyrebase Configuration (use your Firebase project's configuration)
firebaseConfig = {
    "apiKey": "AIzaSyA_S_qNq5UlIaq5Mp2Xl2LdXeOYrgVA51k",
    "authDomain": "your-auth-domain",
    "databaseURL": "your-database-url",
    "projectId": "timesheet-conversion-platform",
    "storageBucket": "your-storage-bucket",
    "messagingSenderId": "865705745377",
    "appId": "your-app-id",
    "measurementId": "your-measurement-id"
}

firebase = pyrebase.initialize_app(firebaseConfig)
auth_client = firebase.auth()
# App State
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False

# Authentication Flow
if st.session_state.logged_in:
    st.title("Timesheet Processing Platform")
    st.write("Logged in successful!")


    # Add any additional features or content here
    ##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # st.header("Timesheet processing Automation platform")
    st.markdown(
        "<h2 style='font-size: 27px; color: white;'>Convert Apexon client employee time sheets(excel, csv file) to Nest standard format.</h2>",
        unsafe_allow_html=True)
    ##++++++++++++++++++++++++++++++++++++Creating Folder structures++++++++++++++++++++++++++++++++++++++++++++++++

    Download_FolderPath = 'Download Folder'
    Individual_FolderPath = 'IndividualEmpFileFolder'
    Empmaster_FolderPath = 'Employee master'
    # Input_dataFolder = 'InputDataFolder'
    Output_resultFolder = 'OutputResultFolder'
    AirBus_LogFolder = 'AirBus LogFolder'
    ICON_LogFolder = 'ICON LogFolder'
    Ford_LogFolder = 'Ford LogFolder'

    # Create the directory if it doesn't exist
    if not os.path.exists(Download_FolderPath):
        os.makedirs(Download_FolderPath)
    if not os.path.exists(Individual_FolderPath):
        os.makedirs(Individual_FolderPath)
    if not os.path.exists(Empmaster_FolderPath):
        os.makedirs(Empmaster_FolderPath)
    # if not os.path.exists(Input_dataFolder):
    #     os.makedirs(Input_dataFolder)
    if not os.path.exists(Output_resultFolder):
        os.makedirs(Output_resultFolder)
    if not os.path.exists(AirBus_LogFolder):
        os.makedirs(AirBus_LogFolder)
    if not os.path.exists(ICON_LogFolder):
        os.makedirs(ICON_LogFolder)
    if not os.path.exists(Ford_LogFolder):
        os.makedirs(Ford_LogFolder)

    ##+++++++++++++++++++Resetting button to clear Files from folder+++++++++++++++++++++++++++
    Reset_files = st.sidebar.button("Click here to clear folders")
    # Output_FormattedButton = st.sidebar.button('Format Output')

    if Reset_files:
        # +++++++++++++++++++++++++++++++++++++++++++Deleting previous files from Output folder++++++++++++++++++++++++++++++++++++++++++++

        # def delete_OutputFolder_files(folder_path, file_list):
        #     for file_name in file_list:
        #         file_path = os.path.join(folder_path, file_name)
        #         if os.path.exists(file_path):
        #             os.remove(file_path)
        #             print(f"Deleted: {file_path}")
        #         else:
        #             print(f"File not found: {file_path}")
        #
        #
        # # Example usage
        # folder_path = Download_FolderPath
        # file_list = ['AirBus_result.csv', 'Ford_result.csv', 'ICON_result.csv']
        # delete_OutputFolder_files(folder_path, file_list)


        # +++++++++++++++++++++++++++++++++++++++++++++++++++++++Deleting previous files from Log folder++++++++++++++++++++++++++++++++++++++++++++++++++++++++
        def delete_LogFolder_files(folder_path, file_list):
            for file_name in file_list:
                file_path = os.path.join(folder_path, file_name)
                if os.path.exists(file_path):
                    os.remove(file_path)
                    print(f"Deleted: {file_path}")
                else:
                    print(f"File not found: {file_path}")


        # Example usage
        folder_path = 'LogFolder'
        file_list = ['AirBusLog_file.csv', 'ICONLog_file.csv', 'Ford_logFile.csv']
        delete_LogFolder_files(folder_path, file_list)


        # ++++++++++++++++++++++++++++++++++++++++++++++++++++++++Deleting previous files from Individual folder+++++++++++++++++++++++++++++++++++++++++++++++++++++++

        def delete_Individual_files(folder_path):
            # Patterns for CSV and Excel files
            patterns = ["*.csv", "*.xlsx", "*.xls"]

            for pattern in patterns:
                # Get a list of all files matching the pattern
                files = glob.glob(os.path.join(folder_path, pattern))

                for file_path in files:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                        print(f"Deleted: {file_path}")
                    else:
                        print(f"File not found: {file_path}")


        # Example usage
        folder_path = 'IndividualEmpFileFolder'
        delete_Individual_files(folder_path)
        st.sidebar.success("Folder has been cleared")

    ##+++++++++++++++++++++++++++++++++++++++++++++++++AirBus process starts++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

    ## AIRBUS CLIENT TIMESHEET

    # Define the paths for bridge files and log file
    airbus_bridge_file = 'Employee master/AirBus bridge file.xlsx'

    # Sidebar select box for project selection
    project = st.selectbox("Select Project", ("--Select--", "AirBus", "FORD", "ICON"))
    if project== 'AirBus':
        # Path to the Excel file
        # Define the provided data in a list of dictionaries
        data = [
            {"SlNo": 1, "EmpID": 201568, "EmployeeName": "Madhusudhanan Kandhababu", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "SP0030C4"},
            {"SlNo": 2, "EmpID": 201136, "EmployeeName": "Surya Prakash C K S", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "NG9407D"},
            {"SlNo": 3, "EmpID": 200952, "EmployeeName": "Chetla Ramesh", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "NG90B63"},
            {"SlNo": 4, "EmpID": 201441, "EmployeeName": "Deepak U S", "ProjectName": "Airbus", "ClientName": "Airbus",
             "AirBus LoginID": "SP002657"},
            {"SlNo": 5, "EmpID": 201079, "EmployeeName": "Harish Y Acharya", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "NG92CC1"},
            {"SlNo": 6, "EmpID": 201360, "EmployeeName": "Iftikhar Ahmed Khan", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "SP001D3E"},
            {"SlNo": 7, "EmpID": 201080, "EmployeeName": "Nagamani Jayapal", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "NG92D4A"},
            {"SlNo": 8, "EmpID": 201435, "EmployeeName": "Padma Praveen Lankalapalli", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "SP002511"},
            {"SlNo": 9, "EmpID": 201268, "EmployeeName": "Ramaprasadreddy Sura", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "SP000D28"},
            {"SlNo": 10, "EmpID": 200961, "EmployeeName": "Shilpa H S", "ProjectName": "Airbus", "ClientName": "Airbus",
             "AirBus LoginID": "NG90FFF"},
            {"SlNo": 11, "EmpID": 201851, "EmployeeName": "Amrita Manjari Pradhan", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "AMANS2UC"},
            {"SlNo": 12, "EmpID": 202769, "EmployeeName": "Poornima H M", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "PHMM7M6G"},
            {"SlNo": 13, "EmpID": 202905, "EmployeeName": "P Sai Suresh", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "SPPPFCR1"},
            {"SlNo": 14, "EmpID": 202992, "EmployeeName": "Tirumani Shanmuka Srilakshmi", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "TSRI2Y6O"},
            {"SlNo": 15, "EmpID": 203147, "EmployeeName": "Sagar Choubey", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "SCHO1XVD"},
            {"SlNo": 16, "EmpID": 203155, "EmployeeName": "Suraj PS", "ProjectName": "Airbus", "ClientName": "Airbus",
             "AirBus LoginID": "SPSSINMS"},
            {"SlNo": 17, "EmpID": 203229, "EmployeeName": "Ajayverma MC", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "AVER5X69"},
            {"SlNo": 18, "EmpID": 203246, "EmployeeName": "Hemamalini Sivaji", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "HSIVVA1E"},
            {"SlNo": 19, "EmpID": 203256, "EmployeeName": "Vishrudh Bav", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "VBAV9YSR"},
            {"SlNo": 20, "EmpID": 203321, "EmployeeName": "Rajashekar NY", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "RNYYM7DP"},
            {"SlNo": 21, "EmpID": 203521, "EmployeeName": "Shalini Priya", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "SPRINI6H"},
            {"SlNo": 22, "EmpID": 203627, "EmployeeName": "Bhagath Chandu Vaddepally", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "BVAD6ZZ9"},
            {"SlNo": 23, "EmpID": 203687, "EmployeeName": "Niranjan Suhas Pimpalwadkar", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "NPIM1XZF"},
            {"SlNo": 24, "EmpID": 203685, "EmployeeName": "Amrutha C", "ProjectName": "Airbus", "ClientName": "Airbus",
             "AirBus LoginID": "ACCC5UA9"},
            {"SlNo": 25, "EmpID": 203834, "EmployeeName": "Reshma Alla", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "RALL5UC4"},
            {"SlNo": 26, "EmpID": 203870, "EmployeeName": "Durga Prasad N", "ProjectName": "Airbus",
             "ClientName": "Airbus", "AirBus LoginID": "DNNNNZ34"}
        ]

        # Convert the data to a Pandas DataFrame
        df = pd.DataFrame(data)


        # Function to create the Excel file in memory
        def create_excel_file(df):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='Airbus Data')
            output.seek(0)  # Rewind the buffer to the beginning so it can be read
            return output


        # Streamlit App
        # st.title("Airbus Employee Data")

        # Display the DataFrame in the Streamlit app
        # st.write("Here is the employee data:")
        # st.dataframe(df)

        # Provide a download button to download the Excel file
        st.download_button(
            label="Download Airbus Bridge File",
            data=create_excel_file(df),
            file_name="Airbus_EmpBridgeFile.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
# _______________________________________________________________________________________________________
    if project== 'FORD':
        # Path to the Excel file
        # Sample data
        data = {
            'SlNo': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20],
            'EmpID': [200325, 200719, 200755, 'NA', 201486, 'NA', 200883, 200890, 204584, 500097, 201033, 500185,
                      500116, 500105, 201187, 'NA', 204983, 204577, 204642, 204774],
            'EmployeeName': ['DENIS AKHMETOV', 'GAYATRI MEHTA', 'VIJAY PAMIDI', 'SHRUTHI SATHISHA', 'SHRUTI KAKIREDDY',
                             'RAMAKRISHNA CHAVA', 'KUSUMA S MYSORE', 'MANASWINI SWAIN', 'DIVYA EDARA',
                             'PREETAM DEVARASETTI', 'TEJASWAROOP MUPPALLA', 'SUDHEER KONDURU',
                             'SAI CHAITANYA MARAPALLI', 'SRINATHBABU KARNATI', 'SRAVANTHI GADE', 'KANDIKANTI TEJESH',
                             'NEERAJ BANGA', 'JONNA SRINIVAS DHARANEESH', 'Deshmukh Sulochana', 'Rondi Praveen'],
            'ProjectName': ['FORD MAGNIT'] * 20,
            'ClientName': ['MAGNIT CORPORATION'] * 20,
            'File Id': [199665, 141379, 215205, 258572, 182623, 247119, 258573, 258526, 258449, 249355, 249992, 251937,
                        252298, 207549, 268350, 272522, 273946, 290903, 291650, 293443]
        }

        # Create a DataFrame
        df = pd.DataFrame(data)


        # Function to create an Excel file in memory
        def create_excel_file(df):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
            output.seek(0)  # Rewind the buffer to the beginning so it can be read
            return output


        # Streamlit app
        # st.title("Download Excel File Directly from App")

        # Display the DataFrame
        # st.write("Here is the data that will be downloaded:")
        # st.dataframe(df)

        # Create the Excel file in memory and provide a download button
        st.download_button(
            label="Download FORD Bridge File",
            data=create_excel_file(df),
            file_name="FORD_EmpBridgeFile.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    #_______________________________________________________________________________________________________
    if project== 'ICON':
        # Path to the Excel file
        # Provided data in a dictionary format
        data = {
            'SlNo': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16],
            'EmpID': [201568, 202425, 202731, 204009, 203762, 204154, 203824, 203851, 203847, 204000, 203131, 203049,
                      203222, 202992, 200430, 203155],
            'EmployeeName': ['MANIKANDAN PUSHKALA', 'JYOTIRMAYEE SATAPATHY', 'NEDUNURI VASANTH KUMAR',
                             'NETHAJI SUBASH CHANDRA BOSE S',
                             'SRIRANJANI.S', 'NISHA YADAV', 'RAJESH THANUMALAYAN', 'NIJANTHAN AV',
                             'CHENCHU VENKATA SIVAPRASAD SAJJA',
                             'RAM KUMAR ADHICHAKRAVARTHY', 'MEGHANA R', 'KETHAVATH PRAKASH', 'SRINIVASAN T',
                             'NEELAKANDAN RAGHUPATHY',
                             'JOHN WESLEY ARE', 'SUGHANYA RAMU'],
            'ProjectName': ['IT-ICON Clinical Research Limited'] * 16,
            'ClientName': ['ICON Clinical Research Limited'] * 16,
            'ICON LoginID': ['NA'] * 16
        }

        # Convert the data to a pandas DataFrame
        df = pd.DataFrame(data)


        # Function to create Excel file in memory
        def create_excel_file(df):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='EmployeeData')
            output.seek(0)  # Rewind the buffer to the beginning so it can be read
            return output


        # Streamlit app
        # st.title("Download Employee Data as Excel")

        # st.write("Here is the data that you can download:")

        # Display the DataFrame in the app
        # st.dataframe(df)

        # Create the Excel file in memory and provide a download button
        st.download_button(
            label="Download ICON Bridge File",
            data=create_excel_file(df),
            file_name="ICON_EmpBridgeFile.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
# _______________________________________________________________________________________________________


    project_bridge_file = st.file_uploader("Upload project Bridge file")

    # Define Path
    AirBus_log_file_path = 'AirBus LogFolder/AirBusLog_file.csv'
    airbus_bridge_file = project_bridge_file


    # if project== 'AirBus':
    #     Initialize log file
    def initialize_log():
        with open(AirBus_log_file_path, 'w') as log_file:
            log_file.write('Log File\n\n')


    # Function to update log file
    def update_log(message):
        with open(AirBus_log_file_path, 'a') as log_file:
            log_file.write(message + '\n')


    # Function to process timesheets

    def process_timesheets(bridge_df, project_name, input_file):
        result_df = pd.DataFrame(
            columns=['SlNo', 'Project Name', 'Date_of_Work', 'Hours', 'Description', 'Is_Billable',
                     'Unique Employee ID'])

        for uploaded_file in input_file:
            if uploaded_file is not None:
                workbook = openpyxl.load_workbook(BytesIO(uploaded_file.read()), data_only=True)
                timesheet_found = False
                employee_found = False

                for sheetname in workbook.sheetnames:
                    sheet = workbook[sheetname]

                    if sheet['A6'].value == 'Name':
                        timesheet_found = True
                        employee_name = sheet['B6'].value
                        # if employee_name is not None and employee_name != '':
                        #     employee_name = employee_name.upper()

                        employee_row = bridge_df[bridge_df['EmployeeName'] == employee_name]
                        if not employee_row.empty:
                            employee_found = True
                            emp_id = employee_row['EmpID'].values[0]

                            if sheet['C11'].value == 'Date' and sheet['D11'].value == 'Hours':

                                temp_df = pd.DataFrame(
                                    columns=['SlNo', 'Project Name', 'Date_of_Work', 'Hours', 'Description',
                                             'Is_Billable',
                                             'Unique Employee ID'])
                                for row in range(12, 42):
                                    work_date = sheet[f'C{row}'].value
                                    hours = sheet[f'D{row}'].value
                                    if work_date and hours:
                                        temp_df = pd.concat([temp_df, pd.DataFrame([{
                                            'SlNo': len(result_df) + len(temp_df) + 1,
                                            'Project Name': project_name,
                                            'Date_of_Work': work_date,
                                            'Hours': hours,
                                            'Description': 'Approved',
                                            'Is_Billable': 'YES',
                                            'Unique Employee ID': emp_id
                                        }])], ignore_index=True)
                                        # # Ensure 'Date_of_Work' column is in datetime format
                                        # temp_df['Date_of_Work'] = pd.to_datetime(temp_df['Date_of_Work'], errors='coerce')
                                        #
                                        # # Convert the 'Date_of_Work' column to the desired format
                                        # temp_df['Date_of_Work'] = temp_df['Date_of_Work'].dt.strftime("%d-%m-%Y")
                                result_df = pd.concat([result_df, temp_df], ignore_index=True)
                            else:
                                update_log(
                                    f'{uploaded_file.name}: C11 or D11 does not contain the expected headers.')
                        else:
                            update_log(
                                f'{uploaded_file.name}: Employee {employee_name} not found in {project_name} bridge file.')

                        break

                if not timesheet_found:
                    update_log(f'{uploaded_file.name}: Timesheet not found.')
                elif not employee_found:
                    update_log(f'{uploaded_file.name}: Employee not found.')

        return result_df


    # Main function to execute the app
    def main():
        initialize_log()
        if project == "AirBus" and airbus_bridge_file:

            input_file = st.file_uploader("Upload AirBus project employee timesheet files.", type='XLSX',
                                          accept_multiple_files=True)
            airbus_bridge_df = pd.read_excel(airbus_bridge_file)
            result_df = process_timesheets(airbus_bridge_df, "Airbus", input_file)

            # Save result to CSV
            result_csv_path = f'Download Folder/AirBus_result_{timestamp}.csv'
            AirBus_resultdf=result_df.to_csv(encoding='utf-8', index=False)
            DownloadResult = result_df.to_csv(result_csv_path, encoding='utf-8', index=False)

            AirBus_result_download = st.download_button(
                label="Generate AirBus result file",
                data=AirBus_resultdf,
                file_name='AirBus_result.csv',
                mime='text/csv'
            )
            if AirBus_result_download:
                st.success("AirBus result file has been downloaded")
            #_______________________________________________________


            #________________________________________________________________

            # Display log and download link
            AirBus_log_download = st.download_button(
                label="Download AirBus log file",
                data=open(AirBus_log_file_path, 'rb').read(),
                file_name='AirBusLog_file.csv',
                mime='text/csv'
            )
            if AirBus_log_download:
                st.success("AirBus log file has been downloaded")
        else:
            pass


    if __name__ == "__main__":
        main()

    ##+++++++++++++++++++++++++++++++++++++++++++++++++++++++AirBus Process Ends++++++++++++++++++++++++++++++++++++++++++++++++++++++

    # ++++++++++++++++++++++++++++++++++++++++++++ICON process starts++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

    ## ICON CLIENT TIMESHEET

    # Define the paths for bridge files and log file
    ICON_bridge_file = project_bridge_file
    ICON_log_file_path = 'ICON LogFolder/ICONLog_file.csv'


    # Initialize log file
    def initialize_log():
        with open(ICON_log_file_path, 'w') as log_file:
            log_file.write('Log File\n\n')


    # Function to update log file
    def update_log(message):
        with open(ICON_log_file_path, 'a') as log_file:
            log_file.write(message + '\n')


    # Function to process timesheets

    def process_timesheets(bridge_df, project_name, input_file):
        result_df = pd.DataFrame(
            columns=['SlNo', 'Project Name', 'Date_of_Work', 'Hours', 'Description', 'Is_Billable',
                     'Unique Employee ID'])

        for uploaded_file in input_file:
            if uploaded_file is not None:
                workbook = openpyxl.load_workbook(BytesIO(uploaded_file.read()), data_only=True)
                timesheet_found = False
                employee_found = False

                for sheetname in workbook.sheetnames:
                    sheet = workbook[sheetname]

                    if sheet['C1'].value == 'Date':
                        timesheet_found = True
                        employee_name = sheet['A2'].value
                        if employee_name is not None and employee_name != "":
                            employee_name = employee_name.upper()

                        employee_row = bridge_df[bridge_df['EmployeeName'] == employee_name]
                        if not employee_row.empty:
                            employee_found = True
                            emp_id = employee_row['EmpID'].values[0]

                            if sheet['C1'].value == 'Date' and sheet['D1'].value == 'Status':
                                temp_df = pd.DataFrame(
                                    columns=['SlNo', 'Project Name', 'Date_of_Work', 'Hours', 'Description',
                                             'Is_Billable',
                                             'Unique Employee ID'])
                                for row in range(2, 32):
                                    work_date = sheet[f'C{row}'].value
                                    status = sheet[f'D{row}'].value
                                    status = status.upper()
                                    if status == 'Present'.upper():
                                        temp_df = pd.concat([temp_df, pd.DataFrame([{
                                            'SlNo': len(result_df) + len(temp_df) + 1,
                                            'Project Name': project_name,
                                            'Date_of_Work': work_date,
                                            'Hours': 8,
                                            'Description': 'Approved',
                                            'Is_Billable': 'YES',
                                            'Unique Employee ID': emp_id
                                        }])], ignore_index=True)
                                result_df = pd.concat([result_df, temp_df], ignore_index=True)
                            else:
                                update_log(f'{uploaded_file.name}: C1 or D1 does not contain the expected headers.')
                        else:
                            update_log(
                                f'{uploaded_file.name}: Employee {employee_name} not found in {project_name} bridge file.')

                        break

                if not timesheet_found:
                    update_log(f'{uploaded_file.name}: Timesheet not found or invalid timesheet format')
                elif not employee_found:
                    update_log(f'{uploaded_file.name}: Employee not found as per the employee master file')

        return result_df


    # Main function to execute the app
    def main():
        initialize_log()
        if project == "ICON" and ICON_bridge_file:

            input_file = st.file_uploader("Upload ICON project employee timesheet files.", type='XLSX',
                                          accept_multiple_files=True)
            ICON_bridge_df = pd.read_excel(ICON_bridge_file)
            result_df = process_timesheets(ICON_bridge_df, "ICON", input_file)
            result_csv_path = f'Download Folder/ICON_result_{timestamp}.csv'
            ICON_resultdf=result_df.to_csv(encoding='utf-8', index=False)
            DownloadResult = result_df.to_csv(result_csv_path, encoding='utf-8', index=False)
            ICON_result_download = st.download_button(
                label="Generate ICON result file",
                data=ICON_resultdf,
                file_name='ICON_result.csv',
                mime='text/csv'
            )
            if ICON_result_download:
                st.success("ICON result file has been downloaded")

            # Display log and download link
            ICON_log_Download = st.download_button(
                label="Download ICON log file",
                data=open(ICON_log_file_path, 'rb').read(),
                file_name='ICONLog_file.csv',
                mime='text/csv'
            )
            if ICON_log_Download:
                st.success("ICON log file has been downloaded")

    if __name__ == "__main__":
        main()
    # +++++++++++++++++++++++++++++++++++++++++++++++++++ICON process Ends++++++++++++++++++++++++++++++++++++++++++++

    ##+++++++++++++++++++++++++++++++++++++++++++++++++++++++Ford Process Starts++++++++++++++++++++++++++++++++++++++++++++++++++++++

    if project == 'FORD':


        timesheet_files = st.file_uploader("Upload Timesheet Files", type="csv",
                                           accept_multiple_files=True)

        if project_bridge_file and timesheet_files:
            try:
                ford_bridge_df = pd.read_excel(project_bridge_file)
                result_df_list = []
                log_result_df_list = []

                for file in timesheet_files:
                    df = pd.read_csv(file, skiprows=1)
                    # Ensure 'Date_of_Work' column is in datetime format
                    df['Work Date'] = pd.to_datetime(df['Work Date'], errors='coerce')

                    # Convert the 'Date_of_Work' column to the desired format
                    df['Work Date'] = df['Work Date'].dt.strftime("%d-%m-%Y")

                    if df.columns[0] == 'PO Number':

                        merged_df = pd.merge(df, ford_bridge_df[['File Id', 'EmpID']], on='File Id', how='inner')
                        filtered_df = merged_df[
                            (merged_df['Work Task'] != 'NBILL') &
                            (merged_df['Work Date'].notna()) &
                            (merged_df['Hours'].notna()) &
                            (merged_df['Hours'] != 0) &
                            (merged_df['Timesheet Status'] != 'MISSING')

                            ]
                        result_df = pd.DataFrame({
                            'SlNo': range(1, len(filtered_df) + 1),
                            'Project Name': 'FORD MAGNIT',
                            'Date_of_Work': filtered_df['Work Date'],
                            'Hours': filtered_df['Hours'],
                            'Description': filtered_df['Timesheet Status'],
                            'Is_Billable': 'YES',
                            'Unique Employee ID': filtered_df['EmpID']
                        })
                        result_df_list.append(result_df)

                        if result_df_list:
                            final_result_df = pd.concat(result_df_list)
                            result_csv_path = f'Download Folder\Ford_result_{timestamp}.csv'
                            ford_resultdf=final_result_df.to_csv(encoding='utf-8', index=False)
                            DownloadResult= final_result_df.to_csv(result_csv_path,encoding='utf-8', index=False)

                Ford_result_Download = st.download_button(label= "Generate Ford result file",
                                                          data= ford_resultdf,
                                                          file_name='Ford_result.csv', mime='text/csv')
                if Ford_result_Download:
                    st.success("Ford result file has been downloaded")


                # # Download Ford Individual file
                # if st.button("Download Ford individual employee file"):
                #     if not final_result_df.empty:
                #         output_dir = 'IndividualEmpFileFolder'
                #         os.makedirs(output_dir, exist_ok=True)
                #         for emp_id in final_result_df['Unique Employee ID'].unique():
                #             emp_df = final_result_df[final_result_df['Unique Employee ID'] == emp_id]
                #             emp_file_path = os.path.join(output_dir, f"{emp_id}.csv")
                #             emp_df.to_csv(emp_file_path, index=False)
                #         st.success(f"Individual employee files created in {output_dir}")
                    ##++++++++++++++++++++++++++++++++++++++++++++++++++++++FORD Log File+++++++++++++++++++++++++++++++++++++++++++
                    if df.columns[0] == 'PO Number':
                        log_merged_df = pd.merge(df, ford_bridge_df[['File Id', 'EmpID']], on='File Id',
                                                 how='inner')
                        logged_df = log_merged_df[
                            (log_merged_df['Work Task'] == 'NBILL') |
                            (log_merged_df['Work Date'].isna()) |
                            (log_merged_df['Hours'].isna()) |
                            (log_merged_df['Hours'] == 0) |
                            (log_merged_df['Timesheet Status'] == 'MISSING')]

                        logresult_df = pd.DataFrame({
                            'SlNo': range(1, len(logged_df) + 1),
                            'Project Name': 'FORD MAGNIT',
                            'Date_of_Work': logged_df['Work Date'],
                            'Hours': logged_df['Hours'],
                            'Description': logged_df['Timesheet Status'],
                            'Is_Billable': 'YES',
                            'Unique Employee ID': logged_df['EmpID'],
                            'Work Task': logged_df['Work Task']
                        })
                        log_result_df_list.append(logresult_df)
                        if log_result_df_list:
                            logfinal_result_df = pd.concat(log_result_df_list)
                            Ford_logFile_path = 'Ford LogFolder\Ford_logFile.csv'
                            ford_logdf=logfinal_result_df.to_csv(encoding='utf-8', index=False)
                            st.download_button(label="Download Ford log file",
                                               data= ford_logdf,
                                               file_name='Ford_logFile.csv', mime='text/csv')



            except Exception as e:
                ("Please select valid timesheet")


    ##++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    if st.sidebar.button("Sign out"):
        st.session_state.logged_in = False
        # st.success('Sign Out successfully')


else:
    # choice = st.selectbox("Login/SignUp", ['Login', 'SignUp'])
    choice = st.selectbox("Login", ['Login'])

    st.title(f"{choice} to Timesheet Processing Platform")
    email = st.text_input('Enter email')
    password = st.text_input('Password', type='password')

    if choice == "Login":
        if st.button("Login"):
            try:
                # Verify credentials
                # user = auth.get_user_by_email(email)
                user = auth_client.sign_in_with_email_and_password(email, password)
                st.session_state.logged_in = True
                # st.success("Login successful!")
                # st.balloons()
            except UserNotFoundError:
                st.error("Login failed. User not found. Please check your credentials.")
            except Exception as e:
                st.error(f"Invalid credentials entered")

    # else:  # Sign Up
    #     if st.button("Sign Up"):
    #         try:
    #             # user = auth.create_user(email=email, password=password)
    #             user = auth_client.create_user_with_email_and_password(email, password)
    #             st.success('Account created successfully')
    #             st.markdown('Please log in using your email and password')
    #             st.balloons()
    #         except Exception as e:
    #             st.error(f"Enter valid email address and password")

##+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

